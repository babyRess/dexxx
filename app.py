from flask import Flask, request, render_template, send_file
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows  # Import the function
import os

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads/'

@app.route('/')
def upload_file():
    return render_template('upload.html')

@app.route('/upload', methods=['POST'])
def upload_json():
    if 'file' not in request.files:
        return "No file part"
    
    file = request.files['file']
    
    if file.filename == '':
        return "No selected file"
    
    if file and file.filename.endswith('.json'):
        data = json.load(file)
        
        # Extract localization parameters
        localizations = data['parameterGroups']['localizations']['parameters']
        localization_keys = [key for key in localizations.keys() if key.endswith(('_en', '_vn', '_hi', '_bn', '_ne'))]

        # Prepare data for DataFrame
        localization_data = []
        seen_parameters = set()  # Set to track seen parameter names

        for key in localization_keys:
            # Extract the base key (without language suffix)
            base_key = key.rsplit('_', 1)[0]
            
            # Check if the base_key has already been seen
            if base_key in seen_parameters:
                continue  # Skip this iteration if the parameter is a duplicate
            seen_parameters.add(base_key)  # Add the base_key to the set

            # Get the values for each language
            english_value = localizations.get(f"{base_key}_en", {}).get('defaultValue', {}).get('value', '')
            vietnamese_value = localizations.get(f"{base_key}_vn", {}).get('defaultValue', {}).get('value', '')
            hindi_value = localizations.get(f"{base_key}_hi", {}).get('defaultValue', {}).get('value', '')
            bengali_value = localizations.get(f"{base_key}_bn", {}).get('defaultValue', {}).get('value', '')
            nepali_value = localizations.get(f"{base_key}_ne", {}).get('defaultValue', {}).get('value', '')

            # Append to the list, ensuring no merging of cells
            localization_data.append([
                base_key,
                english_value,
                vietnamese_value,
                hindi_value,
                bengali_value,
                nepali_value
            ])

        # Create a DataFrame
        df = pd.DataFrame(localization_data, columns=['Parameter Name', 'English Localization', 'Vietnamese Localization', 'Hindi Localization', 'Bengali Localization', 'Nepali Localization'])

        # Save to Excel
        excel_file = 'localizations_with_formatting.xlsx'
        wb = Workbook()
        ws = wb.active

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Set fixed column width to 70
        for column in ws.columns:
            for cell in column:
                ws.column_dimensions[cell.column_letter].width = 70  # Set width to 70

        # Add borders to all cells
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Custom formatting for the header
        header_font = Font(bold=True, color="FFFFFF")  # White font
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue background

        for cell in ws[1]:  # Assuming the first row is the header
            cell.font = header_font
            cell.fill = header_fill

        # Save the workbook
        wb.save(excel_file)

        # Send the generated Excel file for download
        return send_file(excel_file, as_attachment=True)

    return "Invalid file format. Please upload a JSON file."

if __name__ == '__main__':
       app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=True)